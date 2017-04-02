from openpyxl import load_workbook, Workbook

wb = Workbook()
ws = wb.active

ws['A1'].value = "Pizza"
ws['B1'].value = "Belag"
ws['A2'].value = "Funghi"
ws['B2'].value = "Pilze, Käse, Tomaten"
wb.save("excel/funghi.xlsx")

ws['A2'].value = "Margherita"
ws['B2'].value = "Käse, Tomaten"
wb.save("excel/margherita.xlsx")

ws['A2'].value = "Caprese"
ws['B2'].value = "Basilikum, Mozzarella, Tomaten"
wb.save("excel/caprese.xlsx")

main_wb = Workbook()
main_ws = main_wb.active
for pizza in "funghi", "margherita", "caprese":
    wb = load_workbook("./excel/" + pizza + ".xlsx")
    ws = wb.active
    rows = ws.rows
    for row in rows:
        main_ws.append((cell.value for cell in row))

main_wb.save("./excel/pizzen.xlsx")