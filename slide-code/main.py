from PIL import Image
import urllib.request
from html.parser import HTMLParser
from openpyxl import Workbook, load_workbook

# thumbnails
bilder = ["python-1.jpg", "python-2.jpg", "python-3.jpg"]
for bild in bilder:
    aktuelles_bild = Image.open("../images/" + bild)
    aktuelles_bild.thumbnail((120, 120))
    aktuelles_bild.save("../images/thumbnail-" + bild)

# grayscale
bilder = ["python-1.jpg", "python-2.jpg", "python-3.jpg"]
for bild in bilder:
    aktuelles_bild = Image.open("../images/" + bild)
    aktuelles_bild.convert("L").save("../images/gray-" + bild)

# watermark
bilder = ["python-1.jpg", "python-2.jpg", "python-3.jpg"]
for bild in bilder:
    hauptbild = Image.open("../images/" + bild)
    wasserzeichen = Image.open("../images/watermark.png")
    wasserzeichen.thumbnail((hauptbild.size[0], hauptbild.size[1]))
    hauptbild.paste(wasserzeichen, (0, 0), wasserzeichen)
    hauptbild.save("../images/watermark-" + bild)


# urllib and html parsing

page = urllib.request.urlopen("http://www.andcode.de")
html = page.read()
print(html)


class MyParser(HTMLParser):

    def error(self, message):
        pass

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            for attr in attrs:
                if attr[0] == "href" and attr[1].endswith(".csv"):
                    links.add(attr[1])

links = set()


def get_html_from_url(url):
    page = urllib.request.urlopen(url)
    html = page.read()
    return html.decode("utf8")

page_root = "https://support.spatialkey.com/spatialkey-sample-csv-data/"
html = get_html_from_url(page_root)
parser = MyParser()
parser.feed(html)

for i, link in enumerate(links):
    html = get_html_from_url(link)
    f = open("../html/" + str(i) + ".csv", "w")
    f.write(html)
    f.close()


wb = Workbook()
ws = wb.active
ws['A1'].value = "Pizza"
ws['B1'].value = "Belag"
ws['A2'].value = "Funghi"
ws['B2'].value = "Pilze, Käse, Tomaten"
wb.save("../excel/pizza.xlsx")

wb = load_workbook("../excel/pizza.xlsx")
ws = wb.active
# Zeilenweise auslesen
for cell in ws['1']:
    print(cell.value)

# Spaltenweise auslesen
for cell in ws['A']:
    print(cell.value)

main_wb = Workbook()
main_ws = main_wb.active
wb = load_workbook("../excel/pizza.xlsx")
ws = wb.active
rows = ws.rows
for row in rows:
    main_ws.append((cell.value for cell in row))
